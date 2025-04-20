import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import PyPDF2
import tabula
import re
import os
import json
import base64
import io
import tempfile
from datetime import datetime
from fpdf import FPDF
import requests
from PIL import Image
import openpyxl
from io import BytesIO
import openai
import time

# Set page configuration
st.set_page_config(
    page_title="Healthcare Insurance Data Analyzer",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main {
        padding: 2rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 20px;
        border-radius: 4px 4px 0px 0px;
    }
    .report-section {
        background-color: #f8f9fa;
        padding: 20px;
        border-radius: 5px;
        margin-bottom: 20px;
    }
    .footer {
        margin-top: 50px;
        text-align: center;
        font-size: 12px;
        color: #888;
    }
    .highlight {
        background-color: #e7f3fe;
        border-left: 6px solid #2196F3;
        padding: 10px;
        margin: 15px 0;
    }
    .metrics-container {
        display: flex;
        justify-content: space-between;
        flex-wrap: wrap;
    }
    .metric-card {
        background-color: white;
        border-radius: 5px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        padding: 15px;
        margin: 10px 0;
        width: 48%;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state variables if they don't exist
if 'df' not in st.session_state:
    st.session_state.df = None
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = {}
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = {}
if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []
if 'llm_report' not in st.session_state:
    st.session_state.llm_report = ""
if 'column_mapping' not in st.session_state:
    st.session_state.column_mapping = {}

# Function to extract text from PDF
def extract_text_from_pdf(pdf_file):
    text = ""
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text += page.extract_text() + "\n\n"
        return text
    except Exception as e:
        st.error(f"Error extracting text from PDF: {e}")
        return None

# Function to extract tables from PDF using tabula
def extract_tables_from_pdf(pdf_file):
    try:
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
            temp_pdf.write(pdf_file.read())
            pdf_path = temp_pdf.name
        
        # Reset file pointer
        pdf_file.seek(0)
        
        # Extract tables
        tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
        
        # Clean up
        os.unlink(pdf_path)
        
        return tables
    except Exception as e:
        st.error(f"Error extracting tables from PDF: {e}")
        return []

# Function to extract data from Excel
def extract_data_from_excel(excel_file):
    try:
        df = pd.read_excel(excel_file)
        return df
    except Exception as e:
        st.error(f"Error extracting data from Excel: {e}")
        return None

# Function to clean the data
def clean_data(df):
    # Make a copy to avoid modifying the original
    df_clean = df.copy()
    
    # Handle missing values
    st.info(f"Missing values before cleaning:\n{df_clean.isnull().sum()}")
    
    # For numeric columns, fill with median
    numeric_cols = df_clean.select_dtypes(include=['number']).columns
    for col in numeric_cols:
        df_clean[col] = df_clean[col].fillna(df_clean[col].median())
    
    # For categorical columns, fill with "Unknown"
    categorical_cols = df_clean.select_dtypes(include=['object']).columns
    for col in categorical_cols:
        df_clean[col] = df_clean[col].fillna("Unknown")
    
    # Convert date columns to datetime format
    date_columns = [col for col in df_clean.columns if 'date' in col.lower()]
    for col in date_columns:
        try:
            df_clean[col] = pd.to_datetime(df_clean[col])
        except:
            st.warning(f"Could not convert {col} to datetime")
    
    st.info(f"Missing values after cleaning:\n{df_clean.isnull().sum()}")
    return df_clean

# Function to analyze rejections
def analyze_rejections(df, status_col, amount_col, date_col, category_col=None):
    results = {}
    
    # Ensure the required columns exist
    if status_col not in df.columns:
        st.error(f"Status column '{status_col}' not found in data")
        return None
    if amount_col not in df.columns:
        st.error(f"Amount column '{amount_col}' not found in data")
        return None
    
    # Create rejection flag
    df['is_rejected'] = df[status_col].str.contains('reject|denied|declined', case=False)
    
    # Calculate overall rejection rate
    total_claims = df.shape[0]
    rejected_claims = df[df['is_rejected']].shape[0]
    rejection_rate = (rejected_claims / total_claims) * 100
    
    results['overall'] = {
        'total_claims': total_claims,
        'rejected_claims': rejected_claims,
        'rejection_rate': rejection_rate,
        'total_amount': df[amount_col].sum(),
        'rejected_amount': df[df['is_rejected']][amount_col].sum()
    }
    
    # Analyze rejections by category if provided
    if category_col and category_col in df.columns:
        cat_rejection = df.groupby(category_col).agg(
            total_claims=('is_rejected', 'count'),
            rejected_claims=('is_rejected', 'sum'),
            total_amount=(amount_col, 'sum')
        )
        cat_rejection['rejection_rate'] = (cat_rejection['rejected_claims'] / cat_rejection['total_claims']) * 100
        cat_rejection['rejected_amount'] = df[df['is_rejected']].groupby(category_col)[amount_col].sum()
        
        results['by_category'] = cat_rejection.sort_values('rejection_rate', ascending=False).to_dict()
    
    # Analyze rejections over time if date column provided
    if date_col and date_col in df.columns:
        df['month'] = df[date_col].dt.to_period('M')
        time_rejection = df.groupby('month').agg(
            total_claims=('is_rejected', 'count'),
            rejected_claims=('is_rejected', 'sum'),
            total_amount=(amount_col, 'sum')
        )
        time_rejection['rejection_rate'] = (time_rejection['rejected_claims'] / time_rejection['total_claims']) * 100
        time_rejection['rejected_amount'] = df[df['is_rejected']].groupby('month')[amount_col].sum()
        
        results['by_time'] = {str(k): v for k, v in time_rejection.to_dict('index').items()}
    
    return results

# Function to analyze trends
def analyze_trends(df, date_col, value_col, category_col=None, freq='M'):
    results = {}
    
    # Ensure the required columns exist
    if date_col not in df.columns:
        st.error(f"Date column '{date_col}' not found in data")
        return None
    if value_col not in df.columns:
        st.error(f"Value column '{value_col}' not found in data")
        return None
    
    # Set date as index for time series analysis
    df_trend = df.copy()
    
    # Ensure date column is datetime
    if not pd.api.types.is_datetime64_dtype(df_trend[date_col]):
        try:
            df_trend[date_col] = pd.to_datetime(df_trend[date_col])
        except:
            st.error(f"Could not convert {date_col} to datetime")
            return None
    
    # Create period column for grouping
    if freq == 'M':
        df_trend['period'] = df_trend[date_col].dt.to_period('M')
    elif freq == 'Q':
        df_trend['period'] = df_trend[date_col].dt.to_period('Q')
    elif freq == 'Y':
        df_trend['period'] = df_trend[date_col].dt.to_period('Y')
    else:
        df_trend['period'] = df_trend[date_col].dt.to_period('M')  # Default to monthly
    
    # Analyze overall trends
    overall_trend = df_trend.groupby('period').agg(
        count=(value_col, 'count'),
        sum=(value_col, 'sum'),
        mean=(value_col, 'mean'),
        median=(value_col, 'median')
    )
    
    # Calculate growth rates
    overall_trend['sum_growth'] = overall_trend['sum'].pct_change() * 100
    overall_trend['count_growth'] = overall_trend['count'].pct_change() * 100
    
    results['overall'] = {str(k): v for k, v in overall_trend.to_dict('index').items()}
    
    # Analyze trends by category if provided
    if category_col and category_col in df.columns:
        category_trend = df_trend.groupby(['period', category_col]).agg(
            count=(value_col, 'count'),
            sum=(value_col, 'sum')
        )
        
        results['by_category'] = {
            str(period): group.to_dict() 
            for period, group in category_trend.groupby(level=0)
        }
    
    return results

# Function to perform comparative analysis
def compare_categories(df, category_col, value_col, status_col=None):
    results = {}
    
    # Ensure the required columns exist
    if category_col not in df.columns:
        st.error(f"Category column '{category_col}' not found in data")
        return None
    if value_col not in df.columns:
        st.error(f"Value column '{value_col}' not found in data")
        return None
    
    # Basic comparison by category
    category_stats = df.groupby(category_col).agg(
        count=(value_col, 'count'),
        sum=(value_col, 'sum'),
        mean=(value_col, 'mean'),
        median=(value_col, 'median'),
        min=(value_col, 'min'),
        max=(value_col, 'max')
    ).sort_values('sum', ascending=False)
    
    results['basic_stats'] = category_stats.to_dict()
    
    # Add rejection analysis if status column is provided
    if status_col and status_col in df.columns:
        df['is_rejected'] = df[status_col].str.contains('reject|denied|declined', case=False)
        
        rejection_stats = df.groupby(category_col).agg(
            total_claims=('is_rejected', 'count'),
            rejected_claims=('is_rejected', 'sum')
        )
        rejection_stats['rejection_rate'] = (rejection_stats['rejected_claims'] / rejection_stats['total_claims']) * 100
        
        results['rejection_stats'] = rejection_stats.sort_values('rejection_rate', ascending=False).to_dict()
    
    return results

# Function to create visualizations for rejections
def visualize_rejections(df, status_col, amount_col, date_col=None, category_col=None):
    figures = {}
    
    # Create rejection flag
    df['is_rejected'] = df[status_col].str.contains('reject|denied|declined', case=False)
    
    # Overall rejection rate pie chart
    rejected_count = df['is_rejected'].sum()
    approved_count = df.shape[0] - rejected_count
    
    fig_pie = go.Figure(data=[
        go.Pie(
            labels=['Approved', 'Rejected'],
            values=[approved_count, rejected_count],
            hole=0.4,
            marker_colors=['#3498db', '#e74c3c']
        )
    ])
    fig_pie.update_layout(
        title="Overall Rejection Rate",
        height=400
    )
    figures['overall_pie'] = fig_pie
    
    # Rejection amount comparison
    fig_amount = go.Figure()
    total_amount = df[amount_col].sum()
    rejected_amount = df[df['is_rejected']][amount_col].sum()
    approved_amount = total_amount - rejected_amount
    
    fig_amount.add_trace(go.Bar(
        x=['Approved', 'Rejected'],
        y=[approved_amount, rejected_amount],
        marker_color=['#3498db', '#e74c3c']
    ))
    fig_amount.update_layout(
        title="Amount Comparison: Approved vs Rejected",
        yaxis_title="Total Amount",
        height=400
    )
    figures['amount_comparison'] = fig_amount
    
    # Rejection by category if provided
    if category_col and category_col in df.columns:
        category_rejection = df.groupby(category_col).agg(
            total=('is_rejected', 'count'),
            rejected=('is_rejected', 'sum')
        )
        category_rejection['approved'] = category_rejection['total'] - category_rejection['rejected']
        category_rejection['rejection_rate'] = (category_rejection['rejected'] / category_rejection['total']) * 100
        
        # Sort by rejection rate
        category_rejection = category_rejection.sort_values('rejection_rate', ascending=False)
        
        # Only take top 10 categories for visualization
        if len(category_rejection) > 10:
            category_rejection = category_rejection.head(10)
        
        fig_category = go.Figure()
        fig_category.add_trace(go.Bar(
            x=category_rejection.index,
            y=category_rejection['rejection_rate'],
            marker_color='#e74c3c'
        ))
        fig_category.update_layout(
            title=f"Rejection Rate by {category_col}",
            xaxis_title=category_col,
            yaxis_title="Rejection Rate (%)",
            height=500,
            xaxis={'categoryorder': 'total descending'}
        )
        figures['category_rejection'] = fig_category
    
    # Rejection over time if date column provided
    if date_col and date_col in df.columns:
        df['month'] = df[date_col].dt.to_period('M').astype(str)
        
        time_rejection = df.groupby('month').agg(
            total=('is_rejected', 'count'),
            rejected=('is_rejected', 'sum')
        )
        time_rejection['rejection_rate'] = (time_rejection['rejected'] / time_rejection['total']) * 100
        
        fig_time = go.Figure()
        fig_time.add_trace(go.Scatter(
            x=time_rejection.index,
            y=time_rejection['rejection_rate'],
            mode='lines+markers',
            marker_color='#e74c3c'
        ))
        fig_time.update_layout(
            title="Rejection Rate Over Time",
            xaxis_title="Month",
            yaxis_title="Rejection Rate (%)",
            height=400
        )
        figures['time_rejection'] = fig_time
    
    return figures

# Function to create visualizations for trends
def visualize_trends(df, date_col, value_col, category_col=None):
    figures = {}
    
    # Ensure date column is datetime
    df = df.copy()
    if not pd.api.types.is_datetime64_dtype(df[date_col]):
        df[date_col] = pd.to_datetime(df[date_col])
    
    df['month'] = df[date_col].dt.to_period('M').astype(str)
    
    # Overall trend
    monthly_data = df.groupby('month').agg(
        count=(value_col, 'count'),
        sum=(value_col, 'sum'),
        mean=(value_col, 'mean')
    ).reset_index()
    
    # Create interactive visualization for overall trend
    fig_trend = make_subplots(specs=[[{"secondary_y": True}]])
    
    fig_trend.add_trace(
        go.Bar(
            x=monthly_data['month'],
            y=monthly_data['count'],
            name='Claim Count',
            marker_color='#3498db'
        ),
        secondary_y=False
    )
    
    fig_trend.add_trace(
        go.Scatter(
            x=monthly_data['month'],
            y=monthly_data['sum'],
            name='Total Amount',
            marker_color='#2ecc71',
            mode='lines+markers'
        ),
        secondary_y=True
    )
    
    fig_trend.add_trace(
        go.Scatter(
            x=monthly_data['month'],
            y=monthly_data['mean'],
            name='Average Amount',
            marker_color='#f39c12',
            mode='lines+markers',
            line=dict(dash='dash')
        ),
        secondary_y=True
    )
    
    fig_trend.update_layout(
        title="Monthly Claims Trend Analysis",
        xaxis=dict(title="Month"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=500
    )
    
    fig_trend.update_yaxes(title_text="Claim Count", secondary_y=False)
    fig_trend.update_yaxes(title_text=f"Amount", secondary_y=True)
    
    figures['overall_trend'] = fig_trend
    
    # Trend by category if provided
    if category_col and category_col in df.columns:
        # Get top 5 categories by total value
        top_categories = df.groupby(category_col)[value_col].sum().nlargest(5).index.tolist()
        
        # Filter data for top categories
        df_top = df[df[category_col].isin(top_categories)]
        
        # Create category trend visualization
        fig_cat_trend = go.Figure()
        
        for category in top_categories:
            cat_data = df_top[df_top[category_col] == category]
            monthly_cat = cat_data.groupby('month')[value_col].sum().reset_index()
            
            fig_cat_trend.add_trace(go.Scatter(
                x=monthly_cat['month'],
                y=monthly_cat[value_col],
                mode='lines+markers',
                name=str(category)
            ))
        
        fig_cat_trend.update_layout(
            title=f"Monthly Trends by {category_col}",
            xaxis_title="Month",
            yaxis_title=f"Total {value_col}",
            height=500,
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        
        figures['category_trend'] = fig_cat_trend
    
    return figures

# Function to create comparative visualizations
def visualize_comparison(df, category_col, value_col, status_col=None):
    figures = {}
    
    # Basic comparison by category
    category_data = df.groupby(category_col).agg(
        count=(value_col, 'count'),
        sum=(value_col, 'sum'),
        mean=(value_col, 'mean')
    ).reset_index()
    
    # Sort by total amount
    category_data = category_data.sort_values('sum', ascending=False)
    
    # Only take top 10 categories for visualization
    if len(category_data) > 10:
        category_data = category_data.head(10)
    
    # Create bar chart for total amount by category
    fig_total = go.Figure()
    fig_total.add_trace(go.Bar(
        x=category_data[category_col],
        y=category_data['sum'],
        marker_color='#3498db'
    ))
    fig_total.update_layout(
        title=f"Total Amount by {category_col}",
        xaxis_title=category_col,
        yaxis_title="Total Amount",
        height=500,
        xaxis={'categoryorder': 'total descending'}
    )
    figures['total_by_category'] = fig_total
    
    # Create bar chart for average amount by category
    fig_avg = go.Figure()
    fig_avg.add_trace(go.Bar(
        x=category_data[category_col],
        y=category_data['mean'],
        marker_color='#2ecc71'
    ))
    fig_avg.update_layout(
        title=f"Average Amount by {category_col}",
        xaxis_title=category_col,
        yaxis_title="Average Amount",
        height=500,
        xaxis={'categoryorder': 'total descending'}
    )
    figures['avg_by_category'] = fig_avg
    
    # Add rejection comparison if status column is provided
    if status_col and status_col in df.columns:
        df['is_rejected'] = df[status_col].str.contains('reject|denied|declined', case=False)
        
        rejection_data = df.groupby(category_col).agg(
            total=('is_rejected', 'count'),
            rejected=('is_rejected', 'sum')
        ).reset_index()
        rejection_data['rejection_rate'] = (rejection_data['rejected'] / rejection_data['total']) * 100
        
        # Sort by rejection rate
        rejection_data = rejection_data.sort_values('rejection_rate', ascending=False)
        
        # Only take top 10 categories for visualization
        if len(rejection_data) > 10:
            rejection_data = rejection_data.head(10)
        
        fig_rejection = go.Figure()
        fig_rejection.add_trace(go.Bar(
            x=rejection_data[category_col],
            y=rejection_data['rejection_rate'],
            marker_color='#e74c3c'
        ))
        fig_rejection.update_layout(
            title=f"Rejection Rate by {category_col}",
            xaxis_title=category_col,
            yaxis_title="Rejection Rate (%)",
            height=500,
            xaxis={'categoryorder': 'total descending'}
        )
        figures['rejection_by_category'] = fig_rejection
    
    return figures

# Function to generate a report using OpenAI API
def generate_llm_report(data_summary, analysis_results, charts=None, api_key=None):
    st.info("Generating AI analysis report... This might take a moment.")
    
    if not api_key:
        st.error("OpenAI API key not provided. Please enter an API key in the settings.")
        return "Please provide an OpenAI API key to generate reports."
    
    # Set API key
    openai.api_key = api_key
    
    # Prepare the data summary for the LLM
    prompt = f"""
    You are an expert healthcare insurance data analyst. Based on the following data summary and analysis results, provide a detailed report with:
    
    1. Executive Summary
    2. Key Findings
    3. Trend Analysis
    4. Rejection Analysis
    5. Recommendations for Improvement
    
    Format your response in markdown with clear section headings.
    
    DATA SUMMARY:
    {json.dumps(data_summary, indent=2)}
    
    ANALYSIS RESULTS:
    {json.dumps(analysis_results, indent=2)}
    """
    
    try:
        # Call the OpenAI API
        response = openai.ChatCompletion.create(
            model="gpt-4",  # Use an appropriate model
            messages=[
                {"role": "system", "content": "You are an expert healthcare insurance data analyst providing detailed and actionable insights."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=2000
        )
        
        # Get the generated report
        report = response['choices'][0]['message']['content']
        return report
    
    except Exception as e:
        st.error(f"Error generating report with OpenAI API: {e}")
        return f"Error generating report: {str(e)}"

# Function to create Excel report
def create_excel_report(data_summary, analysis_results, filename="insurance_analysis_report.xlsx"):
    try:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create a sheet for data summary
            summary_df = pd.DataFrame({
                'Metric': data_summary.keys(),
                'Value': data_summary.values()
            })
            summary_df.to_excel(writer, sheet_name='Data Summary', index=False)
            
            # Create sheets for each analysis result
            for analysis_name, analysis_data in analysis_results.items():
                if analysis_name == 'rejection_analysis':
                    # Overall rejection metrics
                    if 'overall' in analysis_data:
                        overall_df = pd.DataFrame({
                            'Metric': analysis_data['overall'].keys(),
                            'Value': analysis_data['overall'].values()
                        })
                        overall_df.to_excel(writer, sheet_name='Rejection Overall', index=False)
                    
                    # Rejection by category
                    if 'by_category' in analysis_data:
                        by_cat_df = pd.DataFrame(analysis_data['by_category'])
                        by_cat_df.to_excel(writer, sheet_name='Rejection by Category')
                    
                    # Rejection by time
                    if 'by_time' in analysis_data:
                        by_time_df = pd.DataFrame({k: v for k, v in analysis_data['by_time'].items()})
                        by_time_df.to_excel(writer, sheet_name='Rejection by Time')
                
                elif analysis_name == 'trend_analysis':
                    # Overall trends
                    if 'overall' in analysis_data:
                        overall_df = pd.DataFrame({k: v for k, v in analysis_data['overall'].items()})
                        overall_df.to_excel(writer, sheet_name='Trends Overall')
                    
                    # Trends by category
                    if 'by_category' in analysis_data:
                        # This is more complex, might need to restructure
                        for period, cat_data in analysis_data['by_category'].items():
                            period_df = pd.DataFrame(cat_data)
                            sheet_name = f'Trend_{period[:7]}'  # Limit sheet name length
                            period_df.to_excel(writer, sheet_name=sheet_name)
                
                elif analysis_name == 'comparative_analysis':
                    # Basic stats
                    if 'basic_stats' in analysis_data:
                        basic_df = pd.DataFrame(analysis_data['basic_stats'])
                        basic_df.to_excel(writer, sheet_name='Category Comparison')
                    
                    # Rejection stats
                    if 'rejection_stats' in analysis_data:
                        reject_df = pd.DataFrame(analysis_data['rejection_stats'])
                        reject_df.to_excel(writer, sheet_name='Cat Rejection Stats')
        
        # Reset pointer and return the Excel file
        output.seek(0)
        return output
    
    except Exception as e:
        st.error(f"Error creating Excel report: {e}")
        return None

# Function to create PDF report
def create_pdf_report(data_summary, analysis_results, llm_report="", figures=None):
    try:
        pdf = FPDF()
        pdf.add_page()
        
        # Set fonts
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Healthcare Insurance Data Analysis Report", ln=True, align='C')
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 10, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        pdf.ln(10)
        
        # Add data summary
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Data Summary", ln=True)
        pdf.set_font("Arial", "", 10)
        
        for key, value in data_summary.items():
            pdf.cell(0, 10, f"{key}: {value}", ln=True)
        pdf.ln(5)
        
        # Add analysis results
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Analysis Results", ln=True)
        
        # Add rejection analysis
        if 'rejection_analysis' in analysis_results:
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "Rejection Analysis", ln=True)
            pdf.set_font("Arial", "", 10)
            
            if 'overall' in analysis_results['rejection_analysis']:
                overall = analysis_results['rejection_analysis']['overall']
                pdf.cell(0, 10, f"Overall rejection rate: {overall['rejection_rate']:.2f}%", ln=True)
                pdf.cell(0, 10, f"Total claims: {overall['total_claims']}", ln=True)
                pdf.cell(0, 10, f"Rejected claims: {overall['rejected_claims']}", ln=True)
                pdf.cell(0, 10, f"Total amount: ${overall['total_amount']:,.2f}", ln=True)
                pdf.cell(0, 10, f"Rejected amount: ${overall['rejected_amount']:,.2f}", ln=True)
            pdf.ln(5)
        
        # Add trend analysis
        if 'trend_analysis' in analysis_results:
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "Trend Analysis", ln=True)
            pdf.set_font("Arial", "", 10)
            
            if 'overall' in analysis_results['trend_analysis']:
                pdf.cell(0, 10, "Monthly trends identified in the data", ln=True)
                # We can't easily represent the detailed trend data in simple text format
                # So just mention it's in the detailed report
                pdf.cell(0, 10, "See full Excel report for detailed trend data", ln=True)
            pdf.ln(5)
        
        # Add comparative analysis
        if 'comparative_analysis' in analysis_results:
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "Comparative Analysis", ln=True)
            pdf.set_font("Arial", "", 10)
            
            if 'basic_stats' in analysis_results['comparative_analysis']:
                pdf.cell(0, 10, "Category comparison completed", ln=True)
                pdf.cell(0, 10, "See full Excel report for detailed comparison data", ln=True)
            pdf.ln(5)
        
        # Add LLM report if available
        if llm_report:
            pdf.add_page()
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "AI-Generated Insights", ln=True)
            pdf.set_font("Arial", "", 10)
            
            # Split the markdown report into lines and add to PDF
            # This is a simple approach; a more sophisticated approach would handle markdown formatting
            for line in llm_report.split('\n'):
                # Detect headings (markdown # syntax)
                if line.startswith('# '):
                    pdf.set_font("Arial", "B", 14)
                    pdf.cell(0, 10, line[2:], ln=True)
                    pdf.set_font("Arial", "", 10)
                elif line.startswith('## '):
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, line[3:], ln=True)
                    pdf.set_font("Arial", "", 10)
                elif line.startswith('### '):
                    pdf.set_font("Arial", "B", 11)
                    pdf.cell(0, 10, line[4:], ln=True)
                    pdf.set_font("Arial", "", 10)
                elif line.strip() == '':
                    pdf.ln(5)  # Add some space for empty lines
                else:
                    # Limit line length to fit on page
                    words = line.split()
                    current_line = ""
                    for word in words:
                        test_line = current_line + " " + word if current_line else word
                        if pdf.get_string_width(test_line) < 180:  # Adjust based on page width
                            current_line = test_line
                        else:
                            pdf.cell(0, 6, current_line, ln=True)
                            current_line = word
                    if current_line:
                        pdf.cell(0, 6, current_line, ln=True)
        
        # Get the PDF as bytes
        pdf_output = pdf.output(dest='S').encode('latin1')
        return BytesIO(pdf_output)
    
    except Exception as e:
        st.error(f"Error creating PDF report: {e}")
        return None

# Function to download a file
def get_download_link(file, filename, text):
    b64 = base64.b64encode(file.getvalue()).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{filename}">{text}</a>'
    return href

# Main application layout
def main():
    # Sidebar for navigation
    st.sidebar.title("Healthcare Insurance Data Analyzer")
    st.sidebar.image("https://img.icons8.com/color/96/000000/health-insurance.png", width=100)
    
    # Navigation menu
    page = st.sidebar.radio("Navigation", ["Upload Data", "Data Exploration", "Analysis", "Visualization", "Reports", "Chat with Data", "Settings"])
    
    # API key input in sidebar under settings collapsible section
    with st.sidebar.expander("API Settings"):
        api_key = st.text_input("Enter OpenAI API Key", type="password")
        if api_key:
            st.success("API Key set!")
    
    if page == "Upload Data":
        st.title("Upload Healthcare Insurance Data")
        
        uploaded_files = st.file_uploader("Upload Excel or PDF files", type=["xlsx", "xls", "csv", "pdf"], accept_multiple_files=True)
        
        if uploaded_files:
            st.session_state.uploaded_files = uploaded_files
            
            for file in uploaded_files:
                st.write(f"Uploaded: {file.name}")
                
                if file.name.endswith(('.xlsx', '.xls', '.csv')):
                    # Process Excel file
                    df = extract_data_from_excel(file)
                    if df is not None:
                        st.session_state.df = df
                        st.session_state.extracted_data[file.name] = {
                            'type': 'excel',
                            'data': df
                        }
                        st.success(f"Successfully extracted data from {file.name} with {df.shape[0]} rows and {df.shape[1]} columns")
                
                elif file.name.endswith('.pdf'):
                    # Process PDF file
                    text = extract_text_from_pdf(file)
                    tables = extract_tables_from_pdf(file)
                    
                    if text or tables:
                        st.session_state.extracted_data[file.name] = {
                            'type': 'pdf',
                            'text': text,
                            'tables': tables
                        }
                        
                        if tables:
                            st.success(f"Successfully extracted {len(tables)} tables from {file.name}")
                            
                            # Display the first table as a preview
                            if len(tables) > 0:
                                st.write("Preview of first table:")
                                st.dataframe(tables[0])
                                
                                # Option to convert PDF table to main dataframe
                                if st.button(f"Use first table from {file.name} as main data"):
                                    st.session_state.df = tables[0]
                                    st.success("Table set as main data for analysis")
                        else:
                            st.warning(f"No tables found in {file.name}, but text was extracted")
            
            # If we have a dataframe, show column mapping options
            if st.session_state.df is not None:
                st.subheader("Column Mapping")
                st.write("Please identify the key columns in your data for analysis:")
                
                cols = st.session_state.df.columns.tolist()
                
                col1, col2 = st.columns(2)
                
                with col1:
                    status_col = st.selectbox("Status Column (contains Approved/Rejected)", 
                                            options=[""] + cols, 
                                            format_func=lambda x: f"{x}" if x else "Select a column")
                    
                    amount_col = st.selectbox("Amount Column", 
                                            options=[""] + cols, 
                                            format_func=lambda x: f"{x}" if x else "Select a column")
                
                with col2:
                    date_col = st.selectbox("Date Column", 
                                          options=[""] + cols, 
                                          format_func=lambda x: f"{x}" if x else "Select a column")
                    
                    category_col = st.selectbox("Category Column (for grouping)", 
                                              options=[""] + cols, 
                                              format_func=lambda x: f"{x}" if x else "Select a column")
                
                if st.button("Save Column Mapping"):
                    # Create mapping
                    mapping = {
                        'status_col': status_col if status_col else None,
                        'amount_col': amount_col if amount_col else None,
                        'date_col': date_col if date_col else None,
                        'category_col': category_col if category_col else None
                    }
                    
                    st.session_state.column_mapping = mapping
                    st.success("Column mapping saved!")
                    st.write("You can now proceed to Data Exploration and Analysis")
    
    elif page == "Data Exploration":
        st.title("Data Exploration")
        
        if st.session_state.df is None:
            st.warning("No data uploaded yet. Please go to the Upload Data page first.")
            return
        
        df = st.session_state.df
        
        # Basic information about the data
        st.subheader("Data Overview")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Rows", df.shape[0])
        with col2:
            st.metric("Columns", df.shape[1])
        with col3:
            st.metric("Missing Values", df.isnull().sum().sum())
        
        # Display a sample of the data
        st.subheader("Data Sample")
        st.dataframe(df.head(10))
        
        # Data cleaning
        st.subheader("Data Cleaning")
        if st.button("Clean Data"):
            clean_df = clean_data(df)
            st.session_state.df = clean_df
            st.success("Data cleaned successfully!")
            st.dataframe(clean_df.head(10))
        
        # Data profiling
        st.subheader("Data Profiling")
        
        # Display column information
        st.write("Column Information:")
        col_info = pd.DataFrame({
            'Column': df.columns,
            'Type': df.dtypes,
            'Missing': df.isnull().sum(),
            'Missing %': (df.isnull().sum() / len(df) * 100).round(2),
            'Unique Values': df.nunique()
        })
        st.dataframe(col_info)
        
        # Display summary statistics for numeric columns
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            st.write("Summary Statistics for Numeric Columns:")
            st.dataframe(df[numeric_cols].describe())
        
        # Display unique values for categorical columns (with limit)
        categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
        if categorical_cols:
            st.write("Categorical Columns:")
            
            # Let user select a column to view unique values
            selected_cat_col = st.selectbox("Select column to view unique values", categorical_cols)
            unique_values = df[selected_cat_col].value_counts().head(20)  # Limit to top 20
            
            st.write(f"Top values for {selected_cat_col}:")
            st.dataframe(unique_values)
            
            # Show distribution plot
            fig = px.bar(unique_values, title=f"Distribution of {selected_cat_col}")
            st.plotly_chart(fig)
    
    elif page == "Analysis":
        st.title("Data Analysis")
        
        if st.session_state.df is None:
            st.warning("No data uploaded yet. Please go to the Upload Data page first.")
            return
        
        if not st.session_state.column_mapping:
            st.warning("Column mapping not set. Please go to the Upload Data page to set column mapping.")
            return
        
        df = st.session_state.df
        mapping = st.session_state.column_mapping
        
        # Analysis options
        st.subheader("Select Analysis to Perform")
        
        analysis_type = st.radio(
            "Analysis Type",
            ["Rejection Analysis", "Trend Analysis", "Comparative Analysis", "All"]
        )
        
        if st.button("Run Analysis"):
            with st.spinner("Running analysis..."):
                # Initialize analysis results dict if not exists
                if 'analysis_results' not in st.session_state:
                    st.session_state.analysis_results = {}
                
                # Reset analysis results for clean run
                st.session_state.analysis_results = {}
                
                # Run the selected analysis
                if analysis_type == "Rejection Analysis" or analysis_type == "All":
                    if mapping['status_col'] and mapping['amount_col']:
                        rejection_results = analyze_rejections(
                            df, 
                            status_col=mapping['status_col'], 
                            amount_col=mapping['amount_col'],
                            date_col=mapping['date_col'],
                            category_col=mapping['category_col']
                        )
                        
                        if rejection_results:
                            st.session_state.analysis_results['rejection_analysis'] = rejection_results
                            st.success("Rejection Analysis completed!")
                    else:
                        st.error("Status column and Amount column are required for Rejection Analysis")
                
                if analysis_type == "Trend Analysis" or analysis_type == "All":
                    if mapping['date_col'] and mapping['amount_col']:
                        trend_results = analyze_trends(
                            df,
                            date_col=mapping['date_col'],
                            value_col=mapping['amount_col'],
                            category_col=mapping['category_col']
                        )
                        
                        if trend_results:
                            st.session_state.analysis_results['trend_analysis'] = trend_results
                            st.success("Trend Analysis completed!")
                    else:
                        st.error("Date column and Amount column are required for Trend Analysis")
                
                if analysis_type == "Comparative Analysis" or analysis_type == "All":
                    if mapping['category_col'] and mapping['amount_col']:
                        comp_results = compare_categories(
                            df,
                            category_col=mapping['category_col'],
                            value_col=mapping['amount_col'],
                            status_col=mapping['status_col']
                        )
                        
                        if comp_results:
                            st.session_state.analysis_results['comparative_analysis'] = comp_results
                            st.success("Comparative Analysis completed!")
                    else:
                        st.error("Category column and Amount column are required for Comparative Analysis")
        
        # Display analysis results
        if st.session_state.analysis_results:
            st.subheader("Analysis Results")
            
            # Display rejection analysis results
            if 'rejection_analysis' in st.session_state.analysis_results:
                with st.expander("Rejection Analysis Results", expanded=True):
                    rejection_results = st.session_state.analysis_results['rejection_analysis']
                    
                    if 'overall' in rejection_results:
                        overall = rejection_results['overall']
                        
                        # Create metrics in columns
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Rejection Rate", f"{overall['rejection_rate']:.2f}%")
                        with col2:
                            st.metric("Total Claims", overall['total_claims'])
                        with col3:
                            st.metric("Rejected Claims", overall['rejected_claims'])
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Total Amount", f"${overall['total_amount']:,.2f}")
                        with col2:
                            st.metric("Rejected Amount", f"${overall['rejected_amount']:,.2f}")
                    
                    if 'by_category' in rejection_results and mapping['category_col']:
                        st.subheader(f"Rejection by {mapping['category_col']}")
                        
                        # Convert dict to dataframe for display
                        by_cat_data = pd.DataFrame(rejection_results['by_category'])
                        st.dataframe(by_cat_data)
                    
                    if 'by_time' in rejection_results and mapping['date_col']:
                        st.subheader("Rejection Over Time")
                        
                        # Convert dict to dataframe for display
                        by_time_data = pd.DataFrame({k: v for k, v in rejection_results['by_time'].items()})
                        st.dataframe(by_time_data.T)  # Transpose for better display
            
            # Display trend analysis results
            if 'trend_analysis' in st.session_state.analysis_results:
                with st.expander("Trend Analysis Results", expanded=True):
                    trend_results = st.session_state.analysis_results['trend_analysis']
                    
                    if 'overall' in trend_results:
                        st.subheader("Overall Trends")
                        
                        # Convert dict to dataframe for display
                        overall_trend = pd.DataFrame({k: v for k, v in trend_results['overall'].items()})
                        st.dataframe(overall_trend.T)  # Transpose for better display
                    
                    if 'by_category' in trend_results and mapping['category_col']:
                        st.subheader(f"Trends by {mapping['category_col']}")
                        
                        # This is complex to display in standard tables
                        st.write("Trends by category are available in the visualization tab")
            
            # Display comparative analysis results
            if 'comparative_analysis' in st.session_state.analysis_results:
                with st.expander("Comparative Analysis Results", expanded=True):
                    comp_results = st.session_state.analysis_results['comparative_analysis']
                    
                    if 'basic_stats' in comp_results:
                        st.subheader(f"Comparison by {mapping['category_col']}")
                        
                        # Convert dict to dataframe for display
                        basic_stats = pd.DataFrame(comp_results['basic_stats'])
                        st.dataframe(basic_stats)
                    
                    if 'rejection_stats' in comp_results and mapping['status_col']:
                        st.subheader(f"Rejection Comparison by {mapping['category_col']}")
                        
                        # Convert dict to dataframe for display
                        rejection_stats = pd.DataFrame(comp_results['rejection_stats'])
                        st.dataframe(rejection_stats)
    
    elif page == "Visualization":
        st.title("Data Visualization")
        
        if st.session_state.df is None:
            st.warning("No data uploaded yet. Please go to the Upload Data page first.")
            return
        
        if not st.session_state.column_mapping:
            st.warning("Column mapping not set. Please go to the Upload Data page to set column mapping.")
            return
        
        if not st.session_state.analysis_results:
            st.warning("No analysis results available. Please run analysis first.")
            return
        
        df = st.session_state.df
        mapping = st.session_state.column_mapping
        
        # Create tabs for different visualizations
        tab1, tab2, tab3 = st.tabs(["Rejection Visualizations", "Trend Visualizations", "Comparative Visualizations"])
        
        with tab1:
            if 'rejection_analysis' in st.session_state.analysis_results:
                st.subheader("Rejection Analysis Visualizations")
                
                # Create rejection visualizations
                rejection_figures = visualize_rejections(
                    df,
                    status_col=mapping['status_col'],
                    amount_col=mapping['amount_col'],
                    date_col=mapping['date_col'],
                    category_col=mapping['category_col']
                )
                
                # Display the visualizations
                if 'overall_pie' in rejection_figures:
                    st.plotly_chart(rejection_figures['overall_pie'])
                
                if 'amount_comparison' in rejection_figures:
                    st.plotly_chart(rejection_figures['amount_comparison'])
                
                if 'category_rejection' in rejection_figures and mapping['category_col']:
                    st.plotly_chart(rejection_figures['category_rejection'])
                
                if 'time_rejection' in rejection_figures and mapping['date_col']:
                    st.plotly_chart(rejection_figures['time_rejection'])
            else:
                st.info("No rejection analysis results available. Please run Rejection Analysis first.")
        
        with tab2:
            if 'trend_analysis' in st.session_state.analysis_results:
                st.subheader("Trend Analysis Visualizations")
                
                # Create trend visualizations
                trend_figures = visualize_trends(
                    df,
                    date_col=mapping['date_col'],
                    value_col=mapping['amount_col'],
                    category_col=mapping['category_col']
                )
                
                # Display the visualizations
                if 'overall_trend' in trend_figures:
                    st.plotly_chart(trend_figures['overall_trend'])
                
                if 'category_trend' in trend_figures and mapping['category_col']:
                    st.plotly_chart(trend_figures['category_trend'])
            else:
                st.info("No trend analysis results available. Please run Trend Analysis first.")
        
        with tab3:
            if 'comparative_analysis' in st.session_state.analysis_results:
                st.subheader("Comparative Analysis Visualizations")
                
                # Create comparative visualizations
                comp_figures = visualize_comparison(
                    df,
                    category_col=mapping['category_col'],
                    value_col=mapping['amount_col'],
                    status_col=mapping['status_col']
                )
                
                # Display the visualizations
                if 'total_by_category' in comp_figures:
                    st.plotly_chart(comp_figures['total_by_category'])
                
                if 'avg_by_category' in comp_figures:
                    st.plotly_chart(comp_figures['avg_by_category'])
                
                if 'rejection_by_category' in comp_figures and mapping['status_col']:
                    st.plotly_chart(comp_figures['rejection_by_category'])
            else:
                st.info("No comparative analysis results available. Please run Comparative Analysis first.")
    
    elif page == "Reports":
        st.title("Analysis Reports")
        
        if st.session_state.df is None:
            st.warning("No data uploaded yet. Please go to the Upload Data page first.")
            return
        
        if not st.session_state.analysis_results:
            st.warning("No analysis results available. Please run analysis first.")
            return
        
        # Create data summary
        df = st.session_state.df
        data_summary = {
            "Total Records": df.shape[0],
            "Time Period": f"{df[st.session_state.column_mapping['date_col']].min()} to {df[st.session_state.column_mapping['date_col']].max()}" if st.session_state.column_mapping['date_col'] in df.columns else "N/A",
            "Total Amount": f"${df[st.session_state.column_mapping['amount_col']].sum():,.2f}" if st.session_state.column_mapping['amount_col'] in df.columns else "N/A",
            "Average Amount": f"${df[st.session_state.column_mapping['amount_col']].mean():,.2f}" if st.session_state.column_mapping['amount_col'] in df.columns else "N/A",
        }
        
        # AI Report Generation
        st.subheader("AI-Generated Analysis Report")
        
        if not api_key:
            st.warning("OpenAI API key not set. Please enter it in the Settings section to generate AI reports.")
        else:
            if st.button("Generate AI Report") or st.session_state.llm_report:
                if not st.session_state.llm_report:
                    report = generate_llm_report(data_summary, st.session_state.analysis_results, api_key=api_key)
                    st.session_state.llm_report = report
                
                st.markdown(st.session_state.llm_report)
        
        # Export options
        st.subheader("Export Report")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("Generate Excel Report"):
                excel_file = create_excel_report(data_summary, st.session_state.analysis_results)
                if excel_file:
                    st.markdown(
                        get_download_link(excel_file, "insurance_analysis_report.xlsx", "Download Excel Report"),
                        unsafe_allow_html=True
                    )
        
        with col2:
            if st.button("Generate PDF Report"):
                pdf_file = create_pdf_report(data_summary, st.session_state.analysis_results, st.session_state.llm_report)
                if pdf_file:
                    st.markdown(
                        get_download_link(pdf_file, "insurance_analysis_report.pdf", "Download PDF Report"),
                        unsafe_allow_html=True
                    )
    
    elif page == "Chat with Data":
        st.title("Chat with Your Data")
        
        if st.session_state.df is None:
            st.warning("No data uploaded yet. Please go to the Upload Data page first.")
            return
        
        if not api_key:
            st.warning("OpenAI API key not set. Please enter it in the Settings section to enable chat functionality.")
            return
        
        # Display chat history
        for message in st.session_state.chat_history:
            with st.chat_message(message["role"]):
                st.markdown(message["content"])
        
        # Chat input
        user_query = st.chat_input("Ask a question about your insurance data...")
        
        if user_query:
            # Add user message to chat history
            st.session_state.chat_history.append({"role": "user", "content": user_query})
            
            # Display user message
            with st.chat_message("user"):
                st.markdown(user_query)
            
            # Generate response
            with st.chat_message("assistant"):
                with st.spinner("Thinking..."):
                    try:
                        # Prepare context for the LLM
                        df_info = f"""
                        The dataset has {st.session_state.df.shape[0]} rows and {st.session_state.df.shape[1]} columns.
                        Columns: {', '.join(st.session_state.df.columns.tolist())}
                        
                        Column mapping:
                        - Status Column: {st.session_state.column_mapping.get('status_col', 'None')}
                        - Amount Column: {st.session_state.column_mapping.get('amount_col', 'None')}
                        - Date Column: {st.session_state.column_mapping.get('date_col', 'None')}
                        - Category Column: {st.session_state.column_mapping.get('category_col', 'None')}
                        
                        Data Statistics:
                        {st.session_state.df.describe().to_string()}
                        """
                        
                        # Add analysis results if available
                        analysis_context = ""
                        if st.session_state.analysis_results:
                            analysis_context = f"Analysis Results:\n{json.dumps(st.session_state.analysis_results, indent=2)}"
                        
                        # Call the OpenAI API
                        response = openai.ChatCompletion.create(
                            model="gpt-4",  # Use an appropriate model
                            messages=[
                                {"role": "system", "content": f"You are an expert healthcare insurance data analyst assistant. You are analyzing a dataset with the following information:\n\n{df_info}\n\n{analysis_context}"},
                                {"role": "user", "content": user_query}
                            ],
                            temperature=0.2,
                            max_tokens=1000
                        )
                        
                        # Get the response
                        ai_response = response['choices'][0]['message']['content']
                        st.markdown(ai_response)
                        
                        # Add assistant message to chat history
                        st.session_state.chat_history.append({"role": "assistant", "content": ai_response})
                        
                    except Exception as e:
                        st.error(f"Error generating response: {e}")
                        st.session_state.chat_history.append({"role": "assistant", "content": f"Error generating response: {str(e)}"})
    
    elif page == "Settings":
        st.title("Settings")
        
        st.subheader("API Configuration")
        st.write("Configure API keys for AI report generation and chat functionality")
        
        new_api_key = st.text_input("OpenAI API Key", type="password", value=api_key if api_key else "")
        if st.button("Save API Key"):
            api_key = new_api_key
            st.success("API Key saved successfully!")
        
        st.subheader("Data Settings")
        
        # Option to reset all data
        if st.button("Reset All Data", type="primary", help="This will clear all uploaded data and analysis results"):
            st.session_state.df = None
            st.session_state.uploaded_files = []
            st.session_state.extracted_data = {}
            st.session_state.analysis_results = {}
            st.session_state.chat_history = []
            st.session_state.llm_report = ""
            st.session_state.column_mapping = {}
            st.success("All data has been reset!")
            
        # Add about section
        st.subheader("About")
        st.write("""
        **Healthcare Insurance Data Analyzer** is a tool designed to help analyze healthcare insurance data, 
        identify trends, and provide actionable insights. It supports both Excel and PDF file formats and offers
        various analysis options including rejection analysis, trend analysis, and comparative analysis.
        
        The tool also includes AI-powered reporting and chat functionality powered by OpenAI's GPT models.
        
        Version 1.0 - Developed by Fadil369
        """)

if __name__ == "__main__":
    main()
